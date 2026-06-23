"""生成 Golden Set Review xlsx — 让 Erika 在表格中填写 accept/reject + note.

设计：
- 输入：data/golden_set/v1.0/ai_annotated_500.csv + review_progress.json
- 输出：data/golden_set/v1.0/golden_set_review.xlsx（覆盖旧版）
- Sheet 结构：
  - 操作说明（如何填）
  - 待 review（按优先级排序，未填的在前）
  - 已 review（可只读参考，含历史填写）
  - 子品类汇总（每子品类 × 评分的分布）

填写规则：
- review_action 列：填 'y' = accept AI 标注；填 'n' = reject AI 标注
- review_note 列：reject 时必填修改建议；accept 时可空

下游：填完后 Erika 跑 scripts/import_review_xlsx.py 把表格回写到 review_progress.json
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
INPUT_CSV = ROOT / "data" / "golden_set" / "v1.0" / "ai_annotated_500.csv"
PROGRESS_JSON = ROOT / "data" / "golden_set" / "v1.0" / "review_progress.json"
OUTPUT_XLSX = ROOT / "data" / "golden_set" / "v1.0" / "golden_set_review.xlsx"


def assign_priority(row: pd.Series) -> str:
    rating = row.get("rating", 0)
    sentiment = row.get("ai_sentiment", "")
    evidence = row.get("ai_evidence_level", "")
    if pd.notna(rating):
        if rating <= 3 and sentiment == "positive":
            return "P0"
        if rating >= 4 and sentiment == "negative":
            return "P0"
    if evidence in ("probable", "uncertain"):
        return "P1"
    return "P2"


COLS = [
    "review_id", "priority", "sub_category", "rating", "asin",
    "title", "content", "ai_sentiment", "ai_aspects",
    "ai_pain_points", "ai_highlights", "ai_evidence_level",
    "review_action", "review_note",
]


def main() -> int:
    df = pd.read_csv(INPUT_CSV)
    df = df[df["annotation_status"] == "ai_pending_review"].copy()
    df["priority"] = df.apply(assign_priority, axis=1)

    progress = json.loads(PROGRESS_JSON.read_text(encoding="utf-8")) if PROGRESS_JSON.exists() else {}
    reviewed_ids = set(progress.get("reviewed_ids", []))
    modifications = progress.get("modifications", {})

    df["review_action"] = df["review_id"].map(
        lambda rid: {"accept": "y", "reject": "n"}.get(modifications.get(rid, {}).get("action", ""), "")
    )
    df["review_note"] = df["review_id"].map(
        lambda rid: modifications.get(rid, {}).get("note", "")
    )
    df["_reviewed"] = df["review_id"].isin(reviewed_ids)

    priority_order = {"P0": 0, "P1": 1, "P2": 2}
    df["_p_order"] = df["priority"].map(priority_order).fillna(99).astype(int)

    pending_df = df[~df["_reviewed"]].sort_values(
        ["_p_order", "rating"], ascending=[True, True]
    )[COLS].reset_index(drop=True)

    reviewed_df = df[df["_reviewed"]].sort_values(
        ["_p_order", "rating"], ascending=[True, True]
    )[COLS].reset_index(drop=True)

    print(f"待 review: {len(pending_df)} 条")
    print(f"  P0: {(pending_df['priority']=='P0').sum()} 条")
    print(f"  P1: {(pending_df['priority']=='P1').sum()} 条")
    print(f"  P2: {(pending_df['priority']=='P2').sum()} 条")
    print(f"已 review: {len(reviewed_df)} 条（参考用，不必再填）")

    instructions = pd.DataFrame({
        "项": [
            "📋 Golden Set Review 操作说明",
            "",
            "1. 主战场 sheet「待review」：从上往下填，已按优先级排好",
            "   - P0 = 评分 vs 情感冲突（最高优先级，必看）",
            "   - P1 = AI 自标低置信（probable / uncertain）",
            "   - P2 = 其余样本（默认 AI 已正确，快速 accept 即可）",
            "",
            "2. review_action 列：必填。从下拉选择",
            "   - y = 接受 AI 标注（无须 note）",
            "   - n = 拒绝 AI 标注（必须在 review_note 写为什么 + 正确答案）",
            "",
            "3. review_note 列：仅 reject 时填",
            "   - 写清楚情感是 positive/negative/neutral",
            "   - 简述哪个 aspect 标错了，应该是什么",
            "",
            "4. 中途随时保存 → 工具会增量读取已填的部分",
            "",
            "5. 不必一次填完。建议每天 50-100 条节奏，约 5-10 天完成",
            "",
            "6. 全部填完后，告诉 Claude：'review 表填完了'",
            "   Claude 会跑 scripts/import_review_xlsx.py，把结果回写到 review_progress.json",
            "   然后跑 python3 -m review_analyzer.eval.run --prompt-version v2.1 看新基线",
            "",
            "📊 当前进度",
            f"已 review：{len(reviewed_df)} 条",
            f"待 review：{len(pending_df)} 条",
            f"  P0 优先：{(pending_df['priority']=='P0').sum()} 条",
            f"  P1 优先：{(pending_df['priority']=='P1').sum()} 条",
            f"  P2 优先：{(pending_df['priority']=='P2').sum()} 条",
            "",
            "🎯 验收目标",
            "覆盖率：500 条（V4-T1 计划），目前差距由本次 review 补足",
            "准确率基线：v2.1 在 38 条上 92.1%，扩到 500 条后视为更稳定基线",
        ],
        "说明": [""] * 33,
    })

    summary_rows = []
    for sub_cat in sorted(pending_df["sub_category"].unique()):
        sub = pending_df[pending_df["sub_category"] == sub_cat]
        for r in sorted(sub["rating"].unique()):
            cnt = ((sub["sub_category"] == sub_cat) & (sub["rating"] == r)).sum()
            summary_rows.append({"子品类": sub_cat, "评分": int(r), "待 review 条数": int(cnt)})
    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="📋操作说明", index=False)
        pending_df.to_excel(writer, sheet_name="待review", index=False)
        reviewed_df.to_excel(writer, sheet_name="已review_参考", index=False)
        summary_df.to_excel(writer, sheet_name="子品类分布", index=False)

    # 美化样式
    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_XLSX)

    # 整体样式
    header_fill = PatternFill(start_color="FF682C", end_color="FF682C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    pending_action_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    accept_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    reject_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    p0_fill = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
    p1_fill = PatternFill(start_color="FEF5E7", end_color="FEF5E7", fill_type="solid")
    border = Border(
        left=Side(border_style="thin", color="DDDDDD"),
        right=Side(border_style="thin", color="DDDDDD"),
        top=Side(border_style="thin", color="DDDDDD"),
        bottom=Side(border_style="thin", color="DDDDDD"),
    )

    def style_data_sheet(ws, df_for_sheet: pd.DataFrame, *, allow_edit: bool):
        # 表头
        for col_idx, _col in enumerate(df_for_sheet.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # 列宽
        widths = {
            "review_id": 14,
            "priority": 9,
            "sub_category": 12,
            "rating": 7,
            "asin": 13,
            "title": 24,
            "content": 60,
            "ai_sentiment": 11,
            "ai_aspects": 50,
            "ai_pain_points": 30,
            "ai_highlights": 25,
            "ai_evidence_level": 12,
            "review_action": 13,
            "review_note": 50,
        }
        for col_idx, col in enumerate(df_for_sheet.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col, 14)

        # 冻结首行
        ws.freeze_panes = "C2"

        # 优先级行底色 + content 自动换行
        for row_idx in range(2, len(df_for_sheet) + 2):
            priority = df_for_sheet.iloc[row_idx - 2].get("priority", "")
            row_fill = p0_fill if priority == "P0" else (p1_fill if priority == "P1" else None)
            for col_idx, col in enumerate(df_for_sheet.columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_fill and col not in ("review_action", "review_note"):
                    cell.fill = row_fill

        if not allow_edit:
            # 已 review sheet：整体置浅灰，提示参考
            grey_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            for row_idx in range(2, len(df_for_sheet) + 2):
                for col_idx in range(1, len(df_for_sheet.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.fill.start_color.rgb in (None, "00000000"):
                        cell.fill = grey_fill
            return

        # 待 review sheet：加 review_action 下拉验证 + 高亮列
        action_col_idx = list(df_for_sheet.columns).index("review_action") + 1
        action_letter = get_column_letter(action_col_idx)
        dv = DataValidation(type="list", formula1='"y,n"', allow_blank=True)
        dv.error = "请输入 y（accept）或 n（reject）"
        dv.errorTitle = "无效输入"
        dv.prompt = "y = accept AI 标注\nn = reject AI 标注（review_note 必填）"
        dv.promptTitle = "Review Action"
        ws.add_data_validation(dv)
        dv.add(f"{action_letter}2:{action_letter}{len(df_for_sheet) + 1}")

        # 行尾整列底色
        for row_idx in range(2, len(df_for_sheet) + 2):
            cell = ws.cell(row=row_idx, column=action_col_idx)
            if not cell.value:
                cell.fill = pending_action_fill

        # 条件格式：y/n 染色
        ws.conditional_formatting.add(
            f"{action_letter}2:{action_letter}{len(df_for_sheet) + 1}",
            CellIsRule(operator="equal", formula=['"y"'], fill=accept_fill),
        )
        ws.conditional_formatting.add(
            f"{action_letter}2:{action_letter}{len(df_for_sheet) + 1}",
            CellIsRule(operator="equal", formula=['"n"'], fill=reject_fill),
        )

    style_data_sheet(wb["待review"], pending_df, allow_edit=True)
    style_data_sheet(wb["已review_参考"], reviewed_df, allow_edit=False)

    # 操作说明 sheet 美化
    ws_inst = wb["📋操作说明"]
    ws_inst.column_dimensions["A"].width = 90
    ws_inst.column_dimensions["B"].width = 30
    for row_idx in range(1, len(instructions) + 2):
        for col_idx in range(1, 3):
            cell = ws_inst.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="FF682C")
    # 隐藏第二列（说明列没用上）
    ws_inst.column_dimensions["B"].hidden = True

    # 子品类分布 sheet
    ws_sum = wb["子品类分布"]
    ws_sum.column_dimensions["A"].width = 16
    ws_sum.column_dimensions["B"].width = 8
    ws_sum.column_dimensions["C"].width = 14
    for col_idx in range(1, 4):
        cell = ws_sum.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # sheet 顺序：操作说明 → 待review → 子品类分布 → 已review_参考
    desired_order = ["📋操作说明", "待review", "子品类分布", "已review_参考"]
    wb._sheets = [wb[name] for name in desired_order if name in wb.sheetnames]

    wb.save(OUTPUT_XLSX)
    print(f"\n✅ 已写入 {OUTPUT_XLSX}")
    print(f"   sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
