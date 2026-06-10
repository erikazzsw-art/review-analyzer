"""V4-T1 Step 3 Step E: Taxonomy YAML review 表生成器.

读取 data/taxonomy/v1.0/{cat}/*.yaml + data/taxonomy/v1.0/{home_yamls}.yaml 共 60 个 YAML,
生成 Excel: data/taxonomy/v1.0/REVIEW_SHEET.xlsx

Sheet 1 「aspect_check」: 每行 = (category, sub_category, aspect_key), Erika 在 action 列填:
    keep / delete / rename_to:<new_key> / merge_to:<other_key>
Sheet 2 「phrase_merge」: 跨品类 TOP500 phrase, Erika 在 synonym_of 列填同义词主词

review 完成后, Erika 用同一份 xlsx 给 apply_review_decisions.py (后续脚本) 重生成 yaml.
"""
from __future__ import annotations
import yaml
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
TAXONOMY_DIR = ROOT / "data" / "taxonomy" / "v1.0"
OUTPUT_PATH = TAXONOMY_DIR / "REVIEW_SHEET.xlsx"


def _collect_yamls() -> dict[str, list[Path]]:
    """返回 {category_slug: [yaml_paths]}.

    家居在 v1.0/ 顶层 (沙发.yaml 等), 其他在 v1.0/{cat}/*.yaml.
    """
    groups: dict[str, list[Path]] = {}
    for p in TAXONOMY_DIR.iterdir():
        if p.is_dir():
            yamls = sorted(p.glob("*.yaml"))
            if yamls:
                groups[p.name] = yamls
        elif p.is_file() and p.suffix == ".yaml" and p.name not in ("REVIEW_SHEET.yaml",):
            groups.setdefault("home", []).append(p)
    if "home" in groups:
        groups["home"].sort()
    return groups


def _format_phrases(phrases: list[dict]) -> str:
    """把 [{phrase: 'easy', count: 82}, ...] 拼成单元格可读字符串."""
    parts = []
    for p in phrases[:10]:
        parts.append(f"{p['phrase']} ({p['count']})")
    return " | ".join(parts)


def build_aspect_sheet(wb: openpyxl.Workbook, groups: dict[str, list[Path]]) -> int:
    ws = wb.active
    ws.title = "aspect_check"

    headers = [
        "category", "sub_category", "aspect_key", "label_zh",
        "total", "positive", "negative", "neutral", "negative_rate",
        "top_phrases (TOP10 with counts)", "sample_review_ids",
        "warning",
        "action (keep/delete/rename_to:X/merge_to:X)",
        "note",
    ]
    ws.append(headers)
    # 表头加粗
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 警告色: 浅红=低频(total<10) / 浅黄=极端 polarity (全正面/全负面 且 total>=10) / 浅蓝=高负面率(>50% 且 total>=20)
    fill_low = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    fill_extreme = PatternFill(start_color="FFF6CC", end_color="FFF6CC", fill_type="solid")
    fill_high_neg = PatternFill(start_color="DCEEFD", end_color="DCEEFD", fill_type="solid")

    row_count = 0
    for cat in sorted(groups):
        for ypath in groups[cat]:
            with ypath.open(encoding="utf-8") as f:
                data = yaml.safe_load(f)
            sub = data.get("sub_category", ypath.stem)
            for a in data.get("aspects", []):
                total = a.get("total", 0)
                pos = a.get("positive_count", 0)
                neg = a.get("negative_count", 0)
                neu = a.get("neutral_count", 0)
                neg_rate = a.get("negative_rate", "")

                warnings = []
                if total < 10:
                    warnings.append("低频")
                if total >= 10 and neg == 0:
                    warnings.append("全正面")
                if total >= 10 and pos == 0:
                    warnings.append("全负面")
                if total >= 20:
                    # negative_rate 是 "28.3%" 格式
                    try:
                        rate_val = float(str(neg_rate).rstrip("%"))
                        if rate_val > 50:
                            warnings.append("高负面")
                    except (ValueError, TypeError):
                        pass

                row = [
                    cat,
                    sub,
                    a.get("key", ""),
                    a.get("label_zh", ""),
                    total, pos, neg, neu,
                    neg_rate,
                    _format_phrases(a.get("top_phrases", [])),
                    ", ".join(a.get("sample_reviews", [])[:5]),
                    " | ".join(warnings),
                    "",  # action 留空
                    "",  # note 留空
                ]
                ws.append(row)
                row_count += 1
                # 着色优先级: 高负面 > 极端 > 低频
                cur_row = ws.max_row
                if "高负面" in warnings:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=cur_row, column=c).fill = fill_high_neg
                elif "全正面" in warnings or "全负面" in warnings:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=cur_row, column=c).fill = fill_extreme
                elif "低频" in warnings:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=cur_row, column=c).fill = fill_low

    # 列宽
    widths = [10, 30, 22, 14, 8, 8, 8, 8, 10, 80, 30, 16, 28, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    # 行高 (top_phrases 长)
    ws.row_dimensions[1].height = 32

    # 加 Excel Table 让 Erika 能直接 filter / sort
    end_col = get_column_letter(len(headers))
    end_row = ws.max_row
    table = Table(displayName="AspectCheck", ref=f"A1:{end_col}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True,
    )
    ws.add_table(table)

    return row_count


def build_phrase_sheet(wb: openpyxl.Workbook, groups: dict[str, list[Path]]) -> int:
    """跨品类 TOP500 phrase 同义词合并表."""
    ws = wb.create_sheet("phrase_merge")
    headers = [
        "phrase", "total_count", "appears_in_subs (top 5)",
        "main_aspects (top 3)",
        "synonym_of (Erika 填合并目标 phrase, 留空=保留)",
        "note",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    counter: Counter = Counter()
    sub_index: dict[str, Counter] = {}
    aspect_index: dict[str, Counter] = {}
    for cat, paths in groups.items():
        for ypath in paths:
            with ypath.open(encoding="utf-8") as f:
                data = yaml.safe_load(f)
            sub = data.get("sub_category", ypath.stem)
            for a in data.get("aspects", []):
                key = a.get("key", "")
                for p in a.get("top_phrases", []):
                    phrase = p.get("phrase", "").strip()
                    cnt = int(p.get("count", 0))
                    if not phrase:
                        continue
                    counter[phrase] += cnt
                    sub_index.setdefault(phrase, Counter())[sub] += cnt
                    aspect_index.setdefault(phrase, Counter())[key] += cnt

    for phrase, total in counter.most_common(500):
        top_subs = sub_index[phrase].most_common(5)
        top_aspects = aspect_index[phrase].most_common(3)
        ws.append([
            phrase,
            total,
            " | ".join(f"{s} ({c})" for s, c in top_subs),
            " | ".join(f"{k} ({c})" for k, c in top_aspects),
            "",
            "",
        ])

    widths = [40, 12, 60, 35, 32, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    end_col = get_column_letter(len(headers))
    end_row = ws.max_row
    table = Table(displayName="PhraseMerge", ref=f"A1:{end_col}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4", showRowStripes=True,
    )
    ws.add_table(table)

    return ws.max_row - 1


def build_readme_sheet(wb: openpyxl.Workbook, n_aspect_rows: int, n_phrase_rows: int) -> None:
    ws = wb.create_sheet("README", 0)
    lines = [
        ("V4-T1 Step 3 Taxonomy Review 工作流", True),
        ("", False),
        (f"生成时间: 由 scripts/build_taxonomy_review_sheet.py 生成", False),
        ("", False),
        ("【你需要做什么】", True),
        ("Sheet 「aspect_check」: 60 个 sub_category 共 " + str(n_aspect_rows) + " 个 aspect 实例", False),
        ("  - 行高亮颜色:", False),
        ("    🔴 浅红 = 低频 (total<10), 建议 delete", False),
        ("    🟡 浅黄 = 极端 polarity (全正面/全负面 且 total>=10), 检查 polarity 是否对", False),
        ("    🔵 浅蓝 = 高负面率 (>50% 且 total>=20), 业务上有意义就 keep, 否则 delete", False),
        ("  - 在 action 列填:", False),
        ("    · keep             保留", False),
        ("    · delete           从 yaml 删除该 aspect", False),
        ("    · rename_to:<key>  改 key 名 (例: rename_to:battery_life)", False),
        ("    · merge_to:<key>   合并到该 sub_category 下另一个 aspect (累加 phrases/counts)", False),
        ("  - 留空 = keep", False),
        ("", False),
        ("Sheet 「phrase_merge」: 跨品类 TOP" + str(n_phrase_rows) + " phrase", False),
        ("  - 在 synonym_of 列填: 想合并到的主词 (例 'easy to use')", False),
        ("  - 留空 = 保留原 phrase", False),
        ("", False),
        ("【建议 review 顺序】", True),
        ("1. aspect_check sheet 用 filter 看 warning 列, 优先处理 高负面/极端 polarity", False),
        ("2. 然后看 低频 行, 决定 delete 还是合并", False),
        ("3. phrase_merge sheet 优先合并 TOP30 高频同义词 (easy to use / Easy to use 等大小写差异)", False),
        ("", False),
        ("【提交方式】", True),
        ("Review 完成后, 直接告诉 Claude '我 review 好了'", False),
        ("我会读 REVIEW_SHEET.xlsx, 按 action / synonym_of 列重新生成所有 yaml", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100


def main() -> None:
    groups = _collect_yamls()
    print(f"扫描到 {sum(len(v) for v in groups.values())} 个 YAML 跨 {len(groups)} 个品类")
    for cat, paths in sorted(groups.items()):
        print(f"  {cat}: {len(paths)} 个")

    wb = openpyxl.Workbook()
    n_aspect_rows = build_aspect_sheet(wb, groups)
    n_phrase_rows = build_phrase_sheet(wb, groups)
    build_readme_sheet(wb, n_aspect_rows, n_phrase_rows)

    wb.save(OUTPUT_PATH)
    print(f"\n[OUT] {OUTPUT_PATH}")
    print(f"   aspect_check: {n_aspect_rows} 行")
    print(f"   phrase_merge: {n_phrase_rows} 行")


if __name__ == "__main__":
    main()
