"""V4-T1 Step 3 Step E: 把 REVIEW_SHEET.xlsx 中的 review 决策应用回 60 个 YAML.

读 data/taxonomy/v1.0/REVIEW_SHEET.xlsx 两个 sheet:
- aspect_check: action 列 (keep / delete / rename_to:X / merge_to:X) + note 列 (审计用)
- phrase_merge: synonym_of 列 (合并目标 phrase)

写出 data/taxonomy/v1.0/{cat}/<sub>.yaml 的 v1.1 版 (覆盖原 yaml).
原 yaml 备份到 data/taxonomy/v1.0/<backup-timestamp>/ 防止 review 出错可回滚.

设计原则:
- delete: 直接从 aspects 列表移除
- rename_to:X: aspect_key 改名 (注意若同 sub 已有 X, 视为 merge_to)
- merge_to:X: total/positive_count/negative_count/neutral_count 累加, top_phrases 合并 + 重新排 TOP10,
  sample_reviews 合并 + 去重保留前 5
- phrase synonym_of: 在每个 yaml 的 top_phrases 中, 把 phrase A 改名成 main_phrase, count 累加合并; TOP10 重新排
- negative_rate 重算
- aspect_count 重算
"""
from __future__ import annotations
import re
import shutil
import yaml
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
TAXONOMY_DIR = ROOT / "data" / "taxonomy" / "v1.0"
REVIEW_PATH = TAXONOMY_DIR / "REVIEW_SHEET.xlsx"


def _collect_yamls() -> dict[str, list[Path]]:
    groups: dict[str, list[Path]] = {}
    for p in TAXONOMY_DIR.iterdir():
        if p.is_dir() and not p.name.startswith("backup-"):
            yamls = sorted(p.glob("*.yaml"))
            if yamls:
                groups[p.name] = yamls
        elif p.is_file() and p.suffix == ".yaml":
            groups.setdefault("home", []).append(p)
    if "home" in groups:
        groups["home"].sort()
    return groups


def _parse_aspect_actions(wb: openpyxl.Workbook) -> dict[tuple[str, str], dict]:
    """解析 aspect_check sheet, 返回 {(category, sub_category, aspect_key): action_info}.

    action_info: {"action": "keep"|"delete"|"rename"|"merge", "target": str|None, "note": str|None}
    """
    ws = wb["aspect_check"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    result: dict[tuple[str, str, str], dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        cat = str(row[idx["category"]]).strip() if row[idx["category"]] else ""
        sub = str(row[idx["sub_category"]]).strip() if row[idx["sub_category"]] else ""
        key = str(row[idx["aspect_key"]]).strip() if row[idx["aspect_key"]] else ""
        if not (cat and sub and key):
            continue

        action_col = "action (keep/delete/rename_to:X/merge_to:X)"
        raw = (row[idx[action_col]] or "").strip().lower()
        note = (row[idx["note"]] or "").strip() if row[idx["note"]] else ""

        if not raw or raw == "keep":
            info = {"action": "keep", "target": None, "note": note}
        elif raw == "delete":
            info = {"action": "delete", "target": None, "note": note}
        elif raw.startswith("rename_to:"):
            target = raw.split(":", 1)[1].strip()
            info = {"action": "rename", "target": target, "note": note}
        elif raw.startswith("merge_to:"):
            target = raw.split(":", 1)[1].strip()
            info = {"action": "merge", "target": target, "note": note}
        else:
            print(f"[WARN] 无法解析 action: ({cat}/{sub}/{key}) action={raw!r}")
            info = {"action": "keep", "target": None, "note": note}

        result[(cat, sub, key)] = info
    return result


def _parse_phrase_synonyms(wb: openpyxl.Workbook) -> dict[str, str]:
    """解析 phrase_merge sheet, 返回 {phrase: main_phrase} 同义词映射."""
    if "phrase_merge" not in wb.sheetnames:
        return {}
    ws = wb["phrase_merge"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    syn_col = next((h for h in headers if h and h.startswith("synonym_of")), None)
    if syn_col is None:
        return {}

    syn: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        phrase = str(row[idx["phrase"]]).strip() if row[idx["phrase"]] else ""
        main = (row[idx[syn_col]] or "").strip() if row[idx[syn_col]] else ""
        if phrase and main and phrase != main:
            syn[phrase] = main
    return syn


def _apply_phrase_synonyms(top_phrases: list[dict], syn_map: dict[str, str]) -> list[dict]:
    """合并 phrases 中的同义词, 重新按 count 排序保留 TOP10."""
    if not syn_map:
        return top_phrases
    merged: OrderedDict[str, int] = OrderedDict()
    for p in top_phrases:
        ph = p["phrase"]
        cnt = int(p["count"])
        canonical = syn_map.get(ph, ph)
        merged[canonical] = merged.get(canonical, 0) + cnt
    items = sorted(merged.items(), key=lambda x: -x[1])[:10]
    return [{"phrase": ph, "count": cnt} for ph, cnt in items]


def _merge_aspect(target: dict, source: dict, syn_map: dict[str, str]) -> dict:
    """把 source aspect 合并到 target, 返回 target (in-place modified)."""
    target["total"] = target.get("total", 0) + source.get("total", 0)
    target["positive_count"] = target.get("positive_count", 0) + source.get("positive_count", 0)
    target["negative_count"] = target.get("negative_count", 0) + source.get("negative_count", 0)
    target["neutral_count"] = target.get("neutral_count", 0) + source.get("neutral_count", 0)

    # phrases 合并 + 同义词处理
    combined: list[dict] = list(target.get("top_phrases", [])) + list(source.get("top_phrases", []))
    target["top_phrases"] = _apply_phrase_synonyms(combined, syn_map)

    # sample_reviews 合并去重保留 5
    seen = set()
    samples = []
    for s in list(target.get("sample_reviews", [])) + list(source.get("sample_reviews", [])):
        if s not in seen:
            samples.append(s)
            seen.add(s)
        if len(samples) >= 5:
            break
    target["sample_reviews"] = samples
    return target


def _recompute_aspect(aspect: dict) -> dict:
    """重算 negative_rate."""
    total = aspect.get("total", 0)
    neg = aspect.get("negative_count", 0)
    aspect["negative_rate"] = f"{(neg / total * 100):.1f}%" if total else "0.0%"
    return aspect


def _process_yaml(
    ypath: Path, cat: str,
    actions: dict[tuple[str, str, str], dict],
    syn_map: dict[str, str],
) -> tuple[int, int, int, int]:
    """处理单个 yaml. 返回 (kept, deleted, renamed, merged) 计数."""
    with ypath.open(encoding="utf-8") as f:
        data = yaml.safe_load(f)

    sub = data.get("sub_category", ypath.stem)
    aspects = data.get("aspects", [])

    # 第一遍: 应用 phrase synonyms (所有 aspect 都需做)
    for a in aspects:
        if a.get("top_phrases"):
            a["top_phrases"] = _apply_phrase_synonyms(a["top_phrases"], syn_map)

    # 第二遍: 应用 action
    by_key: OrderedDict[str, dict] = OrderedDict((a["key"], a) for a in aspects)
    deletes: set[str] = set()
    renames: dict[str, str] = {}  # old_key -> new_key
    merges: list[tuple[str, str]] = []  # (source_key, target_key)

    for key in list(by_key.keys()):
        info = actions.get((cat, sub, key))
        if not info:
            continue
        if info["action"] == "delete":
            deletes.add(key)
        elif info["action"] == "rename":
            target = info["target"]
            if not target:
                continue
            # 若同 sub 已有 target, 退化为 merge
            if target in by_key and target != key:
                merges.append((key, target))
            else:
                renames[key] = target
        elif info["action"] == "merge":
            target = info["target"]
            if target and target in by_key and target != key:
                merges.append((key, target))
            elif target:
                # target 不存在, 视为 rename
                renames[key] = target
            # 合并目标自指或为空: 当 keep

    # 执行 merge (累加到 target, 标记 source 删除)
    for src_key, dst_key in merges:
        if src_key not in by_key or dst_key not in by_key:
            continue
        _merge_aspect(by_key[dst_key], by_key[src_key], syn_map)
        deletes.add(src_key)

    # 执行 delete
    for k in deletes:
        by_key.pop(k, None)

    # 执行 rename
    new_by_key: OrderedDict[str, dict] = OrderedDict()
    for k, v in by_key.items():
        new_k = renames.get(k, k)
        v["key"] = new_k
        # 名字冲突时也合并
        if new_k in new_by_key:
            _merge_aspect(new_by_key[new_k], v, syn_map)
        else:
            new_by_key[new_k] = v
    by_key = new_by_key

    # 重算 negative_rate
    for v in by_key.values():
        _recompute_aspect(v)

    # 'other' 移到最后
    if "other" in by_key:
        other = by_key.pop("other")
        by_key["other"] = other

    # 写回 yaml
    new_aspects = list(by_key.values())
    data["aspects"] = new_aspects
    data["aspect_count"] = len(new_aspects)
    data["reviewed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    _write_yaml(ypath, data)

    return (
        len(by_key) - len(merges) - len(renames),  # kept (近似)
        len(deletes),
        len(renames),
        len(merges),
    )


def _write_yaml(path: Path, data: dict) -> None:
    """按原脚本风格写 yaml (手写, 保留 inline 字典格式)."""
    sub = data.get("sub_category", path.stem)
    lines: list[str] = [
        f"# {sub} Aspect Taxonomy v1.1 (after Erika review)",
        f"# 重生成时间: {data.get('reviewed_at', '')}",
        f"sub_category: {sub}",
        f"aspect_count: {data['aspect_count']}",
    ]
    if "reviewed_at" in data:
        lines.append(f"reviewed_at: '{data['reviewed_at']}'")
    lines.append("aspects:")
    for a in data["aspects"]:
        lines.append(f"  - key: {a['key']}")
        lines.append(f"    label_zh: {a.get('label_zh', '')}")
        lines.append(f"    total: {a['total']}")
        lines.append(f"    positive_count: {a['positive_count']}")
        lines.append(f"    negative_count: {a['negative_count']}")
        lines.append(f"    neutral_count: {a['neutral_count']}")
        lines.append(f"    negative_rate: {a['negative_rate']}")
        lines.append(f"    top_phrases:")
        for p in a.get("top_phrases", []):
            ph = str(p["phrase"]).replace('"', '\\"')
            lines.append(f'      - {{phrase: "{ph}", count: {p["count"]}}}')
        sr = a.get("sample_reviews", [])
        lines.append(f"    sample_reviews: {sr}")
        lines.append("")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    if not REVIEW_PATH.exists():
        print(f"错误: 找不到 {REVIEW_PATH}, 请先跑 build_taxonomy_review_sheet.py")
        return 1

    print(f"读取 {REVIEW_PATH}")
    wb = openpyxl.load_workbook(REVIEW_PATH, data_only=True)
    actions = _parse_aspect_actions(wb)
    syn_map = _parse_phrase_synonyms(wb)
    print(f"  解析到 {len(actions)} 行 aspect 决策")
    n_real_actions = sum(1 for v in actions.values() if v["action"] != "keep")
    print(f"  其中 {n_real_actions} 行有实际操作 (delete/rename/merge), 其余 keep")
    print(f"  解析到 {len(syn_map)} 个 phrase 同义词映射")

    if n_real_actions == 0 and not syn_map:
        print("\n[INFO] 未发现任何 review 决策, 不做任何修改 (xlsx 中 action 列和 synonym_of 列均为空).")
        print("       如果你确实没有要改动的, 直接进 Step F/G.")
        return 0

    # 备份
    backup_dir = TAXONOMY_DIR / f"backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    backup_dir.mkdir(parents=True, exist_ok=True)
    print(f"\n备份原 yaml 到 {backup_dir}")

    groups = _collect_yamls()
    total_kept = total_del = total_ren = total_mer = 0
    for cat, paths in sorted(groups.items()):
        cat_backup = backup_dir / cat
        cat_backup.mkdir(parents=True, exist_ok=True)
        for ypath in paths:
            shutil.copy2(ypath, cat_backup / ypath.name)
            kept, deleted, renamed, merged = _process_yaml(ypath, cat, actions, syn_map)
            total_kept += kept
            total_del += deleted
            total_ren += renamed
            total_mer += merged
            if deleted or renamed or merged:
                print(f"  [{cat}/{ypath.name}] -{deleted} ren+{renamed} mrg+{merged}")

    print()
    print("=" * 60)
    print(f"完成 — 删除 {total_del} | 改名 {total_ren} | 合并 {total_mer}")
    print(f"备份: {backup_dir}")
    print(f"如需回滚: cp -r {backup_dir}/* {TAXONOMY_DIR}/")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
