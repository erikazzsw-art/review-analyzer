from __future__ import annotations

import io
from collections import Counter
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from backend_api.app.deps import get_current_user
from backend_api.app.services.customer_label_v2_frontstage import customer_label_v2_frontstage_flag_from_env
from backend_api.app.services.specific_issue import (
    build_customer_highlight_rows,
    build_specific_issue_rows,
    decorate_comment_customer_labels,
)
from review_analyzer.database import get_comments, get_session_by_id
from review_analyzer.exporter import export_to_xlsx
from review_analyzer.quota import InsufficientCreditsError, credit_consume, quota_check

router = APIRouter(prefix="/analysis", tags=["export"])


@router.get("/sessions/{session_id}/export")
def export_module_xlsx(
    session_id: int,
    module: str = Query(default="user_experience"),
    locale: str = Query(default="zh"),
    current_user: dict = Depends(get_current_user),
) -> StreamingResponse:
    user_id = int(current_user["id"])
    session = get_session_by_id(user_id, session_id)
    if not session:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Session not found.")

    allowed, msg = quota_check(user_id, "excel_export")
    if not allowed:
        raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail=msg)

    comments = get_comments(user_id, session_id=session_id)
    for c in comments:
        c.pop("embedding", None)
    v2_frontstage_flag = customer_label_v2_frontstage_flag_from_env()
    comments = [
        decorate_comment_customer_labels(c, locale=locale, v2_frontstage_flag=v2_frontstage_flag)
        for c in comments
    ]

    output = _build_module_xlsx(module, comments, locale)
    try:
        credit_consume(user_id, 1, "export")
    except InsufficientCreditsError as e:
        raise HTTPException(status_code=402, detail=f"Not enough credits: {e.needed} needed, {e.balance} left") from e

    filename = f"analysis_{session_id}_{module}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/sessions/{session_id}/export/full")
def export_full_xlsx(
    session_id: int,
    current_user: dict = Depends(get_current_user),
) -> StreamingResponse:
    """导出完整 4-sheet XLSX（总览摘要 + 源评论明细 + TOP10 问题 + TOP10 亮点）"""
    user_id = int(current_user["id"])
    session = get_session_by_id(user_id, session_id)
    if not session:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Session not found.")

    allowed, msg = quota_check(user_id, "excel_export")
    if not allowed:
        raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail=msg)

    xlsx_bytes, filename = export_to_xlsx(session_id, user_id)
    try:
        credit_consume(user_id, 1, "export")
    except InsufficientCreditsError as e:
        raise HTTPException(status_code=402, detail=f"Not enough credits: {e.needed} needed, {e.balance} left") from e
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


def _top10_headers(locale: str) -> list[str]:
    if locale == "zh":
        return ["排名", "标签", "Mention Count", "Mention Share", "Representative Evidence"]
    return ["Rank", "Tag", "Mention Count", "Mention Share", "Representative Evidence"]


def _customer_highlight_top10_headers(locale: str) -> list[str]:
    if locale == "zh":
        return [
            "排名",
            "客户亮点",
            "Mention Count",
            "Mention Share",
            "Review Count",
            "Impact Review Share",
            "Representative Evidence",
            "内部维度",
            "Canonical Highlight Key",
            "Aspect Key",
            "Highlight Confidence",
            "Evidence Verified",
            "Cluster Propagated",
            "Legacy Fallback",
        ]
    return [
        "Rank",
        "Customer Label",
        "Mention Count",
        "Mention Share",
        "Review Count",
        "Impact Review Share",
        "Representative Evidence",
        "Internal Aspect",
        "Canonical Highlight Key",
        "Aspect Key",
        "Highlight Confidence",
        "Evidence Verified",
        "Cluster Propagated",
        "Legacy Fallback",
    ]


def _specific_issue_top10_headers(locale: str) -> list[str]:
    if locale == "zh":
        return [
            "排名",
            "客户痛点",
            "Mention Count",
            "Mention Share",
            "Review Count",
            "Impact Review Share",
            "Representative Evidence",
            "内部维度",
            "Canonical Issue Key",
            "Aspect Key",
            "Issue Confidence",
            "Evidence Verified",
            "Cluster Propagated",
            "Legacy Fallback",
        ]
    return [
        "Rank",
        "Customer Issue",
        "Mention Count",
        "Mention Share",
        "Review Count",
        "Impact Review Share",
        "Representative Evidence",
        "Internal Aspect",
        "Canonical Issue Key",
        "Aspect Key",
        "Issue Confidence",
        "Evidence Verified",
        "Cluster Propagated",
        "Legacy Fallback",
    ]


def _build_top10_rows(
    pool_comments: list[dict[str, Any]],
    tag_field: str,
) -> list[list[str | int]]:
    tag_counter: Counter[str] = Counter()
    tag_sources: dict[str, list[str]] = {}

    for c in pool_comments:
        raw_tags = c.get(tag_field, "")
        if not raw_tags:
            continue
        seen: set[str] = set()
        for raw_tag in str(raw_tags).split(","):
            tag = raw_tag.strip()
            if tag and tag not in seen:
                seen.add(tag)
                tag_counter[tag] += 1
                if tag not in tag_sources:
                    tag_sources[tag] = []
                content = str(c.get("content", ""))[:120]
                if len(tag_sources[tag]) < 20:
                    tag_sources[tag].append(content)

    pool_size = len(pool_comments) or 1
    rows: list[list[str | int]] = []
    for rank, (tag, count) in enumerate(tag_counter.most_common(10), 1):
        pct = f"{count / pool_size * 100:.1f}%"
        source_text = " | ".join(tag_sources.get(tag, []))
        rows.append([rank, tag, count, pct, source_text])
    return rows


def _num(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _int_num(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _mention_count(row: dict[str, Any]) -> int:
    return _int_num(row.get("mention_count") if row.get("mention_count") is not None else row.get("count"))


def _review_count(row: dict[str, Any]) -> int:
    return _int_num(row.get("review_count") if row.get("review_count") is not None else row.get("count"))


def _mention_share(row: dict[str, Any]) -> float:
    return _num(row.get("mention_share") if row.get("mention_share") is not None else row.get("pct"))


def _impact_review_share(row: dict[str, Any]) -> float:
    return _num(row.get("impact_review_share") if row.get("impact_review_share") is not None else row.get("pct"))


def _pct_text(value: float) -> str:
    return f"{value:.1f}%"


def _representative_evidence(row: dict[str, Any]) -> str:
    if row.get("cluster_propagated") or not row.get("evidence_verified"):
        return ""
    return " | ".join(str(item) for item in (row.get("evidence_spans") or []) if item)


def _build_specific_issue_top10_rows(
    pool_comments: list[dict[str, Any]],
    locale: str,
) -> list[list[str | int]]:
    rows: list[list[str | int]] = []
    for rank, row in enumerate(build_specific_issue_rows(pool_comments, locale=locale), 1):
        aspect_keys = row.get("aspect_keys")
        aspect_key_text = (
            ", ".join(str(item) for item in aspect_keys if item)
            if isinstance(aspect_keys, list)
            else str(row.get("aspect_key") or "")
        )
        rows.append(
            [
                rank,
                str(row.get("specific_issue") or row.get("tag") or ""),
                _mention_count(row),
                _pct_text(_mention_share(row)),
                _review_count(row),
                _pct_text(_impact_review_share(row)),
                _representative_evidence(row),
                str(row.get("dimension") or ""),
                str(row.get("canonical_issue_key") or ""),
                aspect_key_text,
                str(row.get("issue_confidence") or ""),
                "true" if row.get("evidence_verified") else "false",
                "true" if row.get("cluster_propagated") else "false",
                "true" if row.get("legacy_fallback") else "false",
            ]
        )
    return rows


def _build_customer_highlight_top10_rows(
    pool_comments: list[dict[str, Any]],
    locale: str,
) -> list[list[str | int]]:
    rows: list[list[str | int]] = []
    for rank, row in enumerate(build_customer_highlight_rows(pool_comments, locale=locale), 1):
        aspect_keys = row.get("aspect_keys")
        aspect_key_text = (
            ", ".join(str(item) for item in aspect_keys if item)
            if isinstance(aspect_keys, list)
            else str(row.get("aspect_key") or "")
        )
        rows.append(
            [
                rank,
                str(row.get("customer_highlight") or row.get("tag") or ""),
                _mention_count(row),
                _pct_text(_mention_share(row)),
                _review_count(row),
                _pct_text(_impact_review_share(row)),
                _representative_evidence(row),
                str(row.get("dimension") or ""),
                str(row.get("canonical_highlight_key") or ""),
                aspect_key_text,
                str(row.get("highlight_confidence") or ""),
                "true" if row.get("evidence_verified") else "false",
                "true" if row.get("cluster_propagated") else "false",
                "true" if row.get("legacy_fallback") else "false",
            ]
        )
    return rows


def _build_module_xlsx(module_key: str, comments: list[dict[str, Any]], locale: str) -> io.BytesIO:
    import openpyxl

    wb = openpyxl.Workbook()

    headers = _top10_headers(locale)
    issue_headers = _specific_issue_top10_headers(locale)
    highlight_headers = _customer_highlight_top10_headers(locale)

    if module_key == "user_experience":
        ws_pos = wb.active
        pos_title = "正向反馈 TOP10" if locale == "zh" else "Positive Feedback TOP10"
        ws_pos.title = pos_title
        ws_pos.append(highlight_headers)
        for row in _build_customer_highlight_top10_rows(comments, locale):
            ws_pos.append(row)

        neg_title = "负向反馈 TOP10" if locale == "zh" else "Negative Feedback TOP10"
        ws_neg = wb.create_sheet(title=neg_title)
        ws_neg.append(issue_headers)
        for row in _build_specific_issue_top10_rows(comments, locale):
            ws_neg.append(row)

    elif module_key == "purchase_motives":
        ws = wb.active
        ws.title = "Purchase Motives" if locale == "en" else "消费动机"
        ws.append(highlight_headers)
        for row in _build_customer_highlight_top10_rows(comments, locale):
            ws.append(row)

    elif module_key == "unmet_needs":
        ws = wb.active
        ws.title = "Unmet Needs" if locale == "en" else "未满足的需求"
        ws.append(issue_headers)
        for row in _build_specific_issue_top10_rows(comments, locale):
            ws.append(row)

    elif module_key == "consumer_profile":
        ws_pos = wb.active
        pos_title = "亮点标签 TOP10" if locale == "zh" else "Highlight Tags TOP10"
        ws_pos.title = pos_title
        ws_pos.append(highlight_headers)
        for row in _build_customer_highlight_top10_rows(comments, locale):
            ws_pos.append(row)

        neg_title = "问题标签 TOP10" if locale == "zh" else "Issue Tags TOP10"
        ws_neg = wb.create_sheet(title=neg_title)
        ws_neg.append(issue_headers)
        for row in _build_specific_issue_top10_rows(comments, locale):
            ws_neg.append(row)

    elif module_key == "recommendations":
        from review_analyzer.insight_engine import build_results_insights

        context = {"product_id": "", "version": "", "time_label": "", "workflow_purpose": ""}
        modules = build_results_insights(0, comments, context, locale=locale)
        rec_data = modules.get("recommendations", {})
        ws = wb.active
        ws.title = "Recommendations" if locale == "en" else "综合建议"
        rec_headers = (["#", "建议内容"] if locale == "zh" else ["#", "Recommendation"])
        ws.append(rec_headers)
        for i, row in enumerate((rec_data.get("rows") or []), start=1):
            ws.append([i, str(row.get("detail", ""))])

    else:
        ws = wb.active
        ws.title = module_key
        ws.append(headers)

    # AI Transparency disclaimer (California AI Transparency Act AB 2013)
    ai_sheet = wb.create_sheet(title="AI Notice" if locale == "en" else "AI 标注")
    ai_note = "Analysis powered by AI (OpenAI GPT-4o-mini)" if locale == "en" else "AI 生成分析 · 基于 OpenAI GPT-4o-mini"
    ai_sheet.append([ai_note])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
