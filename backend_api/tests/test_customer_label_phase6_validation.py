from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend_api.app.routes.export import _build_module_xlsx
from backend_api.app.services.specific_issue import (
    CUSTOMER_LABEL_OCCURRENCE_RULESET_VERSION,
    CUSTOMER_LABEL_OCCURRENCE_SCHEMA_VERSION,
    build_customer_highlight_rows,
    build_specific_issue_rows,
    iter_customer_highlight_occurrences,
    iter_specific_issue_occurrences,
)
from review_analyzer.exporter import (
    _build_customer_highlight_top10_data,
    _build_specific_issue_top10_data,
)
from review_analyzer.insight_engine import build_results_insights

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "customer_label_phase6_validation.json"


def _phase6_comments() -> list[dict[str, Any]]:
    payload = json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))
    return payload["comments"]


def _comment(comments: list[dict[str, Any]], comment_id: str) -> dict[str, Any]:
    return next(comment for comment in comments if comment["id"] == comment_id)


def _occurrence(
    *,
    label_type: str,
    canonical: str,
    display: str,
    aspect_key: str,
    evidence: str,
    comment_id: str,
    cluster_propagated: bool = False,
) -> dict[str, object]:
    return {
        "comment_id": comment_id,
        "type": label_type,
        "raw_label": display,
        "canonical_label_key": canonical,
        "display_label_en": display,
        "display_label_zh": display,
        "aspect_key": aspect_key,
        "evidence_span": evidence,
        "evidence_start": -1,
        "evidence_end": -1,
        "confidence": "high",
        "source": "human",
        "source_detail": "phase6_unit_fixture",
        "evidence_verified": True,
        "cluster_propagated": cluster_propagated,
        "schema_version": CUSTOMER_LABEL_OCCURRENCE_SCHEMA_VERSION,
        "ruleset_version": CUSTOMER_LABEL_OCCURRENCE_RULESET_VERSION,
        "display_allowed": True,
    }


def test_phase6_fixed_validation_set_top_metrics_and_samples() -> None:
    comments = _phase6_comments()

    issue_rows = build_specific_issue_rows(comments, locale="en", limit=20)
    highlight_rows = build_customer_highlight_rows(comments, locale="en", limit=20)
    total_issue_mentions = sum(int(row["mention_count"]) for row in issue_rows)

    water = next(row for row in issue_rows if row["canonical_issue_key"] == "water_leaks_through")
    assert water["mention_count"] == 2
    assert water["review_count"] == 2
    assert water["count"] == water["mention_count"]
    assert water["pct"] == water["mention_share"]
    assert water["mention_share"] == round(2 / total_issue_mentions * 100, 1)
    assert water["impact_review_share"] == round(2 / len(comments) * 100, 1)
    assert water["evidence_spans"] == ["water leaking in", "Water came in"]

    mixed_issue = next(row for row in issue_rows if row["canonical_issue_key"] == "missing_accessories")
    mixed_highlight = next(
        row for row in highlight_rows if row["canonical_highlight_key"] == "satisfactory_appearance_workmanship"
    )
    assert mixed_issue["specific_issue"] == "Missing Accessories"
    assert mixed_issue["representative_comments"] == [
        "The phone protector and hanger was missing from the package. "
        "The waders seem to be decent and look forward to trying them out."
    ]
    assert mixed_highlight["customer_highlight"] == "Satisfactory Appearance / Workmanship"
    assert mixed_highlight["evidence_spans"] == ["seem to be decent"]

    legacy = next(row for row in issue_rows if row["canonical_issue_key"] == "zipper_quality")
    assert legacy["legacy_fallback"] is True
    assert legacy["evidence_spans"] == ["zipper broke"]
    assert legacy["representative_comments"] == ["Legacy old session: the zipper broke after one use."]

    assert any(row["canonical_issue_key"] == "missing_parts" for row in issue_rows)
    assert any(row["canonical_issue_key"] == "mascara_clumps" for row in issue_rows)
    assert "Waterproof" not in {row["customer_highlight"] for row in highlight_rows}


def test_phase6_evidence_verification_blocks_bad_evidence_and_cluster_representatives() -> None:
    comments = _phase6_comments()

    comfortable_comment = _comment(comments, "comfortable-to-wear-57-missing-evidence")
    comfortable_occurrences = iter_customer_highlight_occurrences(comfortable_comment, locale="en")
    assert len(comfortable_occurrences) == 1
    assert comfortable_occurrences[0]["evidence_verified"] is False
    assert comfortable_occurrences[0]["verified_evidence"] is False

    highlight_rows = build_customer_highlight_rows(comments, locale="en", limit=20)
    assert all(row["canonical_highlight_key"] != "comfortable_to_wear" for row in highlight_rows)

    cluster_comment = _comment(comments, "cluster-propagated-audit-only")
    cluster_occurrences = iter_specific_issue_occurrences(cluster_comment, locale="en")
    assert len(cluster_occurrences) == 1
    assert cluster_occurrences[0]["evidence_verified"] is True
    assert cluster_occurrences[0]["verified_evidence"] is False
    assert cluster_occurrences[0]["cluster_propagated"] is True

    issue_rows = build_specific_issue_rows(comments, locale="en", limit=20)
    assert all(row["canonical_issue_key"] != "pocket_not_waterproof" for row in issue_rows)


def test_phase6_occurrence_aggregation_uses_distinct_reviews_for_denominators() -> None:
    comments = [
        {
            "id": "duplicate-issue-review",
            "content": "The suit was hot and not breathable. It still felt not breathable while walking.",
            "sentiment": "negative",
            "aspects_json": {
                "customer_label_occurrence_schema_version": CUSTOMER_LABEL_OCCURRENCE_SCHEMA_VERSION,
                "sub_category": "apparel",
                "customer_label_occurrences": [
                    _occurrence(
                        label_type="issue",
                        canonical="not_breathable",
                        display="Not Breathable",
                        aspect_key="breathability",
                        evidence="hot and not breathable",
                        comment_id="duplicate-issue-review",
                    ),
                    _occurrence(
                        label_type="issue",
                        canonical="not_breathable",
                        display="Not Breathable",
                        aspect_key="mobility",
                        evidence="not breathable while walking",
                        comment_id="duplicate-issue-review",
                    ),
                ],
            },
        },
        {
            "id": "zipper-issue-review",
            "content": "The zipper broke after one use.",
            "sentiment": "negative",
            "aspects_json": {
                "customer_label_occurrence_schema_version": CUSTOMER_LABEL_OCCURRENCE_SCHEMA_VERSION,
                "sub_category": "apparel",
                "customer_label_occurrences": [
                    _occurrence(
                        label_type="issue",
                        canonical="zipper_fails",
                        display="Zipper Fails",
                        aspect_key="zipper_quality",
                        evidence="zipper broke",
                        comment_id="zipper-issue-review",
                    )
                ],
            },
        },
    ]

    rows = build_specific_issue_rows(comments, locale="en", limit=10)
    breathable = next(row for row in rows if row["canonical_issue_key"] == "not_breathable")

    assert breathable["raw_occurrence_count"] == 2
    assert breathable["mention_count"] == 1
    assert breathable["review_count"] == 1
    assert breathable["mention_share"] == 50.0
    assert breathable["impact_review_share"] == 50.0


def test_phase6_insight_engine_keeps_mixed_sentiment_occurrences(monkeypatch: Any) -> None:
    monkeypatch.setattr(
        "review_analyzer.insight_engine._build_ai_results_payload",
        lambda *args, **kwargs: None,
    )

    insights = build_results_insights(1, _phase6_comments(), {"product_id": "phase6"}, locale="en")
    issue_keys = {
        row["canonical_issue_key"] for row in insights["user_experience"]["negative"] if row.get("canonical_issue_key")
    }
    highlight_keys = {
        row["canonical_highlight_key"]
        for row in insights["user_experience"]["positive"]
        if row.get("canonical_highlight_key")
    }

    assert "water_leaks_through" in issue_keys
    assert "missing_accessories" in issue_keys
    assert "feels_well_made" in highlight_keys
    assert "satisfactory_appearance_workmanship" in highlight_keys


def test_phase6_export_payload_contains_audit_fields_and_matches_rows() -> None:
    comments = _phase6_comments()
    output = _build_module_xlsx("user_experience", comments, "en")
    workbook = load_workbook(output)
    negative_sheet = workbook["Negative Feedback TOP10"]
    positive_sheet = workbook["Positive Feedback TOP10"]

    negative_headers = [cell.value for cell in negative_sheet[1]]
    positive_headers = [cell.value for cell in positive_sheet[1]]
    for headers in (negative_headers, positive_headers):
        for required in (
            "Mention Count",
            "Mention Share",
            "Review Count",
            "Impact Review Share",
            "Representative Evidence",
            "Evidence Verified",
            "Cluster Propagated",
        ):
            assert required in headers

    neg_rows = list(negative_sheet.iter_rows(min_row=2, values_only=True))
    pos_rows = list(positive_sheet.iter_rows(min_row=2, values_only=True))
    neg_by_label = {str(row[1]): row for row in neg_rows if row[1]}
    pos_by_label = {str(row[1]): row for row in pos_rows if row[1]}

    assert "Pocket Not Waterproof" not in neg_by_label

    assert "Comfortable To Wear" not in pos_by_label

    issue_headers, issue_rows = _build_specific_issue_top10_data(comments)
    highlight_headers, highlight_rows = _build_customer_highlight_top10_data(comments)
    assert "Cluster Propagated" in issue_headers
    assert "Cluster Propagated" in highlight_headers

    assert all(row[1] != "口袋不防水" for row in issue_rows)
