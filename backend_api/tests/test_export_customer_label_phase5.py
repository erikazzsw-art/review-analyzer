from __future__ import annotations

import io

from openpyxl import load_workbook

from backend_api.app.routes.export import _build_module_xlsx
from backend_api.app.services.specific_issue import (
    CUSTOMER_LABEL_OCCURRENCE_RULESET_VERSION,
    CUSTOMER_LABEL_OCCURRENCE_SCHEMA_VERSION,
)
from review_analyzer.exporter import (
    _build_customer_highlight_top10_data,
    _build_specific_issue_top10_data,
    export_to_xlsx,
)


def _occurrence(
    *,
    label_type: str,
    canonical: str,
    display_en: str,
    display_zh: str,
    aspect_key: str,
    evidence: str,
    comment_id: int,
) -> dict[str, object]:
    return {
        "comment_id": comment_id,
        "type": label_type,
        "raw_label": display_en,
        "canonical_label_key": canonical,
        "display_label_en": display_en,
        "display_label_zh": display_zh,
        "aspect_key": aspect_key,
        "evidence_span": evidence,
        "evidence_start": -1,
        "evidence_end": -1,
        "confidence": "high",
        "source": "rule",
        "source_detail": "test_occurrence",
        "evidence_verified": True,
        "cluster_propagated": False,
        "schema_version": CUSTOMER_LABEL_OCCURRENCE_SCHEMA_VERSION,
        "ruleset_version": CUSTOMER_LABEL_OCCURRENCE_RULESET_VERSION,
        "display_allowed": True,
    }


def _mixed_review_comments() -> list[dict[str, object]]:
    return [
        {
            "id": 1,
            "content": "I like the waders, but the zipper broke on day one.",
            "sentiment": "positive",
            "aspects_json": {
                "customer_label_occurrence_schema_version": CUSTOMER_LABEL_OCCURRENCE_SCHEMA_VERSION,
                "sub_category": "outdoor",
                "customer_label_occurrences": [
                    _occurrence(
                        label_type="issue",
                        canonical="zipper_fails",
                        display_en="Zipper Fails",
                        display_zh="拉链故障",
                        aspect_key="zipper_quality",
                        evidence="zipper broke",
                        comment_id=1,
                    )
                ],
            },
        },
        {
            "id": 2,
            "content": "The fit was bad, but they kept me dry.",
            "sentiment": "negative",
            "aspects_json": {
                "customer_label_occurrence_schema_version": CUSTOMER_LABEL_OCCURRENCE_SCHEMA_VERSION,
                "sub_category": "outdoor",
                "customer_label_occurrences": [
                    _occurrence(
                        label_type="highlight",
                        canonical="keeps_water_out",
                        display_en="Keeps Water Out",
                        display_zh="防水可靠",
                        aspect_key="waterproof",
                        evidence="kept me dry",
                        comment_id=2,
                    )
                ],
            },
        },
    ]


def test_module_export_uses_phase4_fields_and_all_comments_for_mixed_reviews() -> None:
    output = _build_module_xlsx("user_experience", _mixed_review_comments(), "en")
    wb = load_workbook(output)

    positive = wb["Positive Feedback TOP10"]
    negative = wb["Negative Feedback TOP10"]

    assert [cell.value for cell in negative[1]][:7] == [
        "Rank",
        "Customer Issue",
        "Mention Count",
        "Mention Share",
        "Review Count",
        "Impact Review Share",
        "Representative Evidence",
    ]
    assert "Sentiment-Pool Mention Share" not in [cell.value for cell in negative[1]]
    assert negative.cell(row=2, column=2).value == "Zipper Fails"
    assert negative.cell(row=2, column=3).value == 1
    assert negative.cell(row=2, column=4).value == "100.0%"
    assert negative.cell(row=2, column=5).value == 1
    assert negative.cell(row=2, column=6).value == "50.0%"
    assert negative.cell(row=2, column=7).value == "zipper broke"

    assert positive.cell(row=2, column=2).value == "Keeps Water Out"
    assert positive.cell(row=2, column=6).value == "50.0%"
    assert positive.cell(row=2, column=7).value == "kept me dry"


def test_full_export_top_data_uses_phase4_headers_and_fields() -> None:
    issue_headers, issue_rows = _build_specific_issue_top10_data(_mixed_review_comments())
    highlight_headers, highlight_rows = _build_customer_highlight_top10_data(_mixed_review_comments())

    assert issue_headers[:7] == [
        "排名",
        "客户痛点",
        "Mention Count",
        "Mention Share",
        "Review Count",
        "Impact Review Share",
        "Representative Evidence",
    ]
    assert all("情绪池" not in header for header in issue_headers + highlight_headers)
    assert issue_rows[0][1] == "拉链故障"
    assert issue_rows[0][2:7] == ["1", "100.0%", "1", "50.0%", "zipper broke"]
    assert highlight_rows[0][1] == "防水可靠"
    assert highlight_rows[0][2:7] == ["1", "100.0%", "1", "50.0%", "kept me dry"]


def test_full_export_generates_valid_xlsx_with_ai_notice_sheet(monkeypatch) -> None:
    from review_analyzer import exporter

    monkeypatch.setattr(
        exporter,
        "get_session_by_id",
        lambda user_id, session_id: {
            "id": session_id,
            "product_id": "Foxelli",
            "version": "phase6",
            "total_reviews": 2,
            "positive_count": 1,
            "negative_count": 1,
            "category": "waders",
            "created_at": "2026-07-25T00:00:00Z",
        },
    )
    monkeypatch.setattr(exporter, "get_comments", lambda user_id, session_id: _mixed_review_comments())

    xlsx_bytes, filename = export_to_xlsx(3, 6)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))

    assert filename.endswith(".xlsx")
    assert "AI Notice" in workbook.sheetnames
    assert "AI Notice / AI 标注" not in workbook.sheetnames
